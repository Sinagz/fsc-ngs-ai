"""In-memory openpyxl workbook generator. Columns derived from the pydantic schema."""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.matching import LookupResult
from src.pipeline.schema import FeeCodeRecord

EXPORT_FIELDS: list[str] = [
    name for name in FeeCodeRecord.model_fields if name != "schema_version"
] + ["sim_score", "ngs_match", "score_method", "is_anchor"]

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_ANCHOR_FILL = PatternFill("solid", fgColor="FFE699")
_PROV_FILLS = {
    "ON": PatternFill("solid", fgColor="DDEEFF"),
    "BC": PatternFill("solid", fgColor="E8F8E8"),
    "YT": PatternFill("solid", fgColor="FFF3DD"),
}
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(vertical="top", wrap_text=False)


def _row(
    record: FeeCodeRecord,
    *,
    sim_score: float | None,
    ngs_match: bool | None,
    score_method: str | None,
    is_anchor: bool,
) -> list:
    dumped = record.model_dump(mode="json")
    base = [
        dumped.get(f, "")
        for f in FeeCodeRecord.model_fields
        if f != "schema_version"
    ]
    base.extend([
        "" if sim_score is None else f"{sim_score:.4f}",
        "" if ngs_match is None else ("yes" if ngs_match else "no"),
        score_method or "",
        "yes" if is_anchor else "no",
    ])
    return base


def build_workbook(results: Iterable[LookupResult]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "FSC Lookup"

    # Header row
    for col_idx, name in enumerate(EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(name) + 2)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    next_row = 2
    for result in results:
        anchor_row = _row(
            result.anchor, sim_score=None, ngs_match=None,
            score_method=result.score_method, is_anchor=True,
        )
        for col_idx, value in enumerate(anchor_row, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.fill = _ANCHOR_FILL
            cell.alignment = _BODY_ALIGN
        next_row += 1

        for match in result.matches:
            match_row = _row(
                match.fee_code, sim_score=match.sim_score,
                ngs_match=match.ngs_match, score_method=match.score_method,
                is_anchor=False,
            )
            for col_idx, value in enumerate(match_row, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.fill = _PROV_FILLS.get(match.fee_code.province, PatternFill())
                cell.alignment = _BODY_ALIGN
            next_row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_FIELDS))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
